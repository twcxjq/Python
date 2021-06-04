# -*- coding:utf8 -*- #
# -----------------------------------------------------------------------------------
# ProjectName:   python_29
# FileName:     py_excel
# Author:      TianChangJun
# Datetime:    2021/6/3 17:12
# Description：
# -----------------------------------------------------------------------------------

'''
from openpyxl import Workbook
import datetime

# 实例化对象:作用是打开excel文件
wb = Workbook()
# 选择sheet页,默认
# get_sheet = wb.active
# 指定sheet页和设置sheet名字
get_sheet = wb.create_sheet("名单", 2)
# get_sheet["A10"] = "张伟"
# get_sheet.append([1,2,3,4])
# # 对单元格进行写入内容
# get_sheet["a6"] = "花儿乐队"
# get_sheet["b11"] = datetime.datetime.now()

data = [
    ['序号', '组号', '申报单位', '姓名', '准考号', '考试分数(卷面分）']
    , ['1', '1', '新民采油厂', '刘铁', '2012051224', '67.834']
    , ['2', '1', '吉林油田总医院', '吕册', '2012120214', '66.776']
    , ['3', '1', '吉林油田总医院', '王彦苏', '2012120718', '66.683']
    , ['4', '1', '农业开发公司', '刘化峰', '2012060120', '66.664']
    , ['5', '1', '油区教育处', '沙丽珊', '2012020817', '66.657']
    , ['6', '1', '油区教育处', '王志会', '2012020914', '66.545']
    , ['7', '1', '扶余采油厂', '姜波', '2012041210', '66.355']
    , ['8', '2', '红岗采油厂', '王昕', '2012021226', '66.3']
    , ['9', '2', '通信公司', '彭莉', '2012121014', '66.17']
    , ['10', '2', '前郭矿区', '米云龙', '2012070121', '65.313']
    , ['11', '2', '油区教育处', '秦勤', '2012022620', '65.061']
    , ['12', '2', '储运销售公司', '许华', '2012130406', '64.95']
    , ['13', '2', '江南物业', '兰庆伟', '2012030620', '64.869']
    , ['14', '2', '滨江物业', '佟冬蕾', '2012110417', '64.652']
    , ['15', '2', '农业开发公司', '曹宽', '2012061025', '64.564']
    , ['16', '2', '农业开发公司', '孙成伟', '2012060318', '64.462']
    , ['17', '3', '红岗采油厂', '张大勇', '2012020112', '64.384']
    , ['18', '3', '储运销售公司', '刘贤宇', '2012130721', '64.378']
    , ['19', '3', '建设公司', '李月', '2012010316', '64.029']
    , ['20', '3', '客运公司', '姚佳媛', '2012130613', '63.914']
    , ['21', '3', '物资供应处', '张丽', '2012110925', '63.883']
    , ['22', '3', '江北物业', '益长虹', '2012100222', '63.723']
    , ['23', '3', '建设公司', '张雁冰', '2012013327', '63.630']
    , ['24', '3', '公用事业管理公司', '王涵', '2012022214', '63.522']
]
for i in data:
    get_sheet.append(i)


# 保存
wb.save(r"c:\test12.xlsx")
'''


"""
# 批量向excel添加数据

from openpyxl import Workbook

# 实例化对象,作用是打开excel文件
wb = Workbook()
# 选择sheet页,默认
# get_sheet = wb.active
# 指定shee页和设置sheet页名字
get_sheet = wb.create_sheet("名单", 2)
# 获取文件路径，并打开文件
with open(r"D:\文件\QQ下载\test.txt") as f1:
    data = f1.read()  # 读取文件内容，内容是字符串
    # print(type(data), data)
    list_data = eval(data)    # 将字符串转换成列表
    for i in list_data:
        get_sheet.append(i)   # append()方法中的参数: Value must be a list, tuple, range or generator, or a dict
# 保存
wb.save("test11.xlsx")
"""

# 批量读取excel文件中的数据

"""
import openpyxl

# 打开文件
wb = openpyxl.load_workbook(r"D:\下载\QQ下载\计科2班软件工程分组名单.xlsx")
# 选择sheet页
get_sheet = wb["分组"]
# 第一个for循环的是一行的数据，以元组的形式存在
for hang_tuple in get_sheet:
    # 第二个for循环的是单元格的内容
    for i in hang_tuple:
        # print(i.row, i.column, i.value)
        # print(i.coordinate, i.value)
        if i.value != None:
            print(i, i.value)
"""

import openpyxl

# 打开文件
wb = openpyxl.load_workbook(r"D:\下载\QQ下载\计科2班软件工程分组名单.xlsx")
# 选择sheet页
get_sheet = wb["分组"]
# 第一个for循环的是一行的数据，以元组的形式存在
all_list = []
for hang_tuple in get_sheet:
    # 第二个for循环的是单元格的内容
    list_22 = []
    for i in hang_tuple:
        # print(i.row, i.column, i.value)
        # print(i.coordinate, i.value)
        if i.value != None:
            # print(i, i.value)
            a = i.value
            # print(a)
            list_22.append(a)
    all_list.append(list_22)
    print(list_22)
print(all_list)





