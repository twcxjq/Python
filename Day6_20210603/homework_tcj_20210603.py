# -*- coding:utf8 -*- #
# -----------------------------------------------------------------------------------
# ProjectName:   python_29
# FileName:     homework_tcj_20210603
# Author:      TianChangJun
# Datetime:    2021/6/3 17:35
# Description：
# -----------------------------------------------------------------------------------

from openpyxl import Workbook
import pymysql

"""
1,把你们组的成员名字添加到excel表中并将代码封装到类中
"""


class TeamMingDan:
    def addMingDan(self, list_team):
        # 实例化Workbook对象
        wb = Workbook()
        # 选择默认的sheet页
        get_sheet = wb.active
        list_team = ["田昌君", "王昌盛", "王婷", "王海涛"]
        get_sheet.append(list_team)
        # 保存当前路径下
        wb.save("team.xlsx")
        return 0


if __name__ == '__main__':
    list_team = ["田昌君", "王昌盛", "王婷", "王海涛"]
    tmd = TeamMingDan()
    tmd.addMingDan(list_team)

"""
2,把您们的名字添加到表tbl_sms_plan中：格式：29期_名字并将代码封装到类中
"""


class NameTable:
    def addNameToTable(self, list_person_message):
        my_connect = pymysql.connect(host="106.13.199.209", user="root", password="ronghuanet@2019",
                                     database="message", port=3306, charset="utf8")
        # 创建游标，作用是便利所有的数据存放到游标中，以便后面操作
        youbiao = my_connect.cursor()
        # 执行SQL语句
        youbiao.execute("INSERT INTO tbl_sms_plan (id,NAME,company_id,TYPE,STATUS,create_dt,update_dt) VALUES ({},'{}',{},{},{},NOW(),NOW())".format(list_person_message[0], list_person_message[1], list_person_message[2], list_person_message[3], list_person_message[4]))
        # 提交数据
        my_connect.commit()
        # 获取执行后的数据
        youbiao.execute("select * from tbl_sms_plan where id=101")
        data = youbiao.fetchall()
        print(data[0])
        # 关闭数据库与游标
        my_connect.close()
        youbiao.close()
        return 0


if __name__ == '__main__':
    list_person_message = [101, "29期_tcj", 1997, 1, 1]
    nt = NameTable()
    nt.addNameToTable(list_person_message)

"""
3,把表tbl_sms_plan中的数据读取出来添加到excel表中并将代码封装到类中
"""


class PlanData:
    def getPlanDataToExcel(self):
        my_connect1 = pymysql.connect(host="106.13.199.209", user="root", password="ronghuanet@2019",
                                      database="message", port=3306, charset="utf8")
        # 创建游标，作用是便利所有的数据存放到游标中，以便后面操作
        youbiao1 = my_connect1.cursor()
        # 执行SQL语句
        youbiao1.execute("select * from tbl_sms_plan")
        data1 = youbiao1.fetchall()
        # print(type(data1), data1)
        # 关闭数据库与游标
        my_connect1.close()
        youbiao1.close()
        wb1 = Workbook()
        get_sheet_1 = wb1.create_sheet("名单", 2)
        for i in data1:
            get_sheet_1.append(i)
        wb1.save("mindan.xlsx")
        return 0


if __name__ == '__main__':
    pd = PlanData()
    pd.getPlanDataToExcel()

"""
1,把你们组的成员名字添加到excel表中
"""
# from openpyxl import Workbook
# import pymysql

# # 实例化Workbook对象
# wb = Workbook()
# # 选择默认的sheet页
# get_sheet = wb.active
# list_team = ["田昌君", "王昌盛", "王婷", "王海涛"]
# get_sheet.append(list_team)
# # 保存当前路径下
# wb.save("team.xlsx")

"""
2,把您们的名字添加到表tbl_sms_plan中：格式：29期_名字
"""
# my_connect = pymysql.connect(host="106.13.199.209", user="root", password="ronghuanet@2019",
#                              database="message", port=3306, charset="utf8")
# #创建游标，作用是便利所有的数据存放到游标中，以便后面操作
# youbiao = my_connect.cursor()
# #执行SQL语句
# youbiao.execute("INSERT INTO tbl_sms_plan (id,NAME,company_id,TYPE,STATUS,create_dt,update_dt) VALUES (100,'29期_田昌君',1997,1,1,NOW(),NOW());")
# #提交数据
# my_connect.commit()
# # 获取执行后的数据
# youbiao.execute("select * from tbl_sms_plan where id=100;")
# data = youbiao.fetchall()
# print(data[0])
# #关闭数据库与游标
# my_connect.close()
# youbiao.close()

"""
3,把表tbl_sms_plan中的数据读取出来添加到excel表中
"""
# my_connect1 = pymysql.connect(host="106.13.199.209", user="root", password="ronghuanet@2019",
#                              database="message", port=3306, charset="utf8")
# #创建游标，作用是便利所有的数据存放到游标中，以便后面操作
# youbiao1 = my_connect1.cursor()
# #执行SQL语句
# youbiao1.execute("select * from tbl_sms_plan")
# data1 = youbiao1.fetchall()
# # print(type(data1), data1)
# #关闭数据库与游标
# my_connect1.close()
# youbiao1.close()
#
# wb1 = Workbook()
# get_sheet_1 = wb1.create_sheet("名单", 2)
# for i in data1:
#     get_sheet_1.append(i)
# wb1.save("mindan.xlsx")
